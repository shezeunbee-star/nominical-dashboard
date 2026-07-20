"""
플랫폼 출고 파일 → 구글 시트 자동 정리
사용법: python3 platform_to_sheets.py [파일경로1] [파일경로2] ...
파일명으로 플랫폼 자동 감지:
  - 29CM_출고관리리스트_*.xlsx → 29CM
  - 상품준비중내역*.xlsx       → W컨셉
  - 배송조회*.xls              → SSF
  - 발송처리목록*.xlsx         → SI Village
  - 최근 주문내역*.xlsx          → Cafe24 (자사몰)
  - 결제완료내역*.xlsx           → Cafe24 (자사몰)
"""
import sys, os, re
import openpyxl, xlrd, gspread
from google.oauth2.service_account import Credentials as SACredentials
from datetime import datetime

SA_FILE        = os.path.join(os.path.dirname(os.path.abspath(__file__)), "service_account.json")
SPREADSHEET_ID = "1y9mZirj81sR2tkkGV_wTzFvJonPdJU-JuErSRDo_73E"
SHEET_NAME     = "🏬 플랫폼 매출"
COMMISSION        = 30  # 공통 수수료율 (%)
COMMISSION_CAFE24 = 3   # Cafe24 자사몰 PG 수수료율 (%)

creds = SACredentials.from_service_account_file(SA_FILE, scopes=[
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
])
gc = gspread.authorize(creds)
ws = gc.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)

HEADERS = ["플랫폼", "주문일", "상품명", "상품코드", "컬러", "사이즈", "수량", "판매가", "수수료율(%)", "실수익", "주문상태"]

def ensure_header():
    first = ws.row_values(1)
    if not first or first[0] != "플랫폼":
        ws.update(values=[HEADERS], range_name="A1:K1")
        print("  헤더 작성 완료")

def get_existing_keys():
    """중복 방지용 기존 데이터 키 (플랫폼+주문일+상품코드+컬러+사이즈+수량) 수집
    - 같은 날 같은 상품을 다른 사람이 주문해도 별도 건으로 처리하기 위해 수량 포함
    """
    all_rows = ws.get_all_values()
    keys = {}  # key → count (같은 키가 몇 번 등장했는지)
    for row in all_rows[1:]:
        if len(row) >= 7:
            k = f"{row[0]}|{row[1]}|{row[3]}|{row[4]}|{row[5]}|{row[6]}"
            keys[k] = keys.get(k, 0) + 1
    return keys

# ── 컬러/사이즈 파서 ─────────────────────────────────────────────

def parse_color_size_29cm(option):
    """[컬러]블랙 / [사이즈]L 형식"""
    color, size = "-", "-"
    if not option:
        return color, size
    m_color = re.search(r'\[컬러\]([^/\[\]]+)', str(option))
    m_size  = re.search(r'\[사이즈\]([^/\[\]]+)', str(option))
    if m_color: color = m_color.group(1).strip()
    if m_size:  size  = m_size.group(1).strip()
    return color, size

def parse_color_size_slash(option):
    """컬러/사이즈 또는 사이즈만 (SSF, SI Village 공통)"""
    color, size = "-", "-"
    if not option or str(option).strip() == "":
        return color, size
    opt = str(option).strip()
    if "/" in opt:
        parts = opt.split("/", 1)
        color = parts[0].strip()
        size  = parts[1].strip()
    elif opt.upper() == "FREE":
        size = "FREE"
    else:
        size = opt
    return color, size

# ── 날짜 포맷터 ─────────────────────────────────────────────────

def fmt_date(val):
    """일반 날짜 → YYYY-MM-DD"""
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    m = re.match(r'(\d{4}-\d{2}-\d{2})', s)
    if m:
        return m.group(1)
    # YYYYMMDD 형식
    if re.match(r'\d{8}$', s[:8]):
        try:
            return datetime.strptime(s[:8], "%Y%m%d").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""

def find_col(header, *keywords):
    """헤더 행에서 키워드가 포함된 컬럼 인덱스 반환 (없으면 None)"""
    for kw in keywords:
        for i, h in enumerate(header):
            if h and kw in str(h).strip():
                return i
    return None

def fmt_date_sivillage(val):
    """SI Village 날짜: '20260525061821' → '2026-05-25'"""
    if not val:
        return ""
    s = str(val).strip()
    if len(s) >= 8 and s.isdigit():
        try:
            return datetime.strptime(s[:8], "%Y%m%d").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return fmt_date(val)

# ── 플랫폼별 파서 ───────────────────────────────────────────────

def parse_29cm(filepath):
    print(f"  📦 29CM 파싱: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath)
    ws_f = wb.active
    all_rows = list(ws_f.iter_rows(values_only=True))
    if not all_rows:
        return []
    header = all_rows[0]
    # 헤더에서 주문일 컬럼 자동 탐지
    i_date   = find_col(header, "주문일", "주문 일", "주문일시") or 28
    i_prod   = find_col(header, "상품명") or 8
    i_code   = find_col(header, "상품코드", "품목코드") or 7
    i_opt    = find_col(header, "옵션", "옵션명") or 10
    i_qty    = find_col(header, "수량") or 11
    i_price  = find_col(header, "판매가", "상품금액", "단가") or 20
    i_status = find_col(header, "주문상태", "처리상태") or 33
    rows = []
    for i, row in enumerate(all_rows):
        if i == 0: continue
        if not row[i_prod]: continue
        color, size = parse_color_size_29cm(row[i_opt])
        qty    = int(row[i_qty]) if row[i_qty] else 1
        price  = int(row[i_price]) if row[i_price] else 0
        total  = price * qty
        profit = round(total * (1 - COMMISSION / 100))
        status = str(row[i_status]) if row[i_status] else ""
        rows.append([
            "29CM", fmt_date(row[i_date]),
            str(row[i_prod]), str(row[i_code]),
            color, size, qty, total,
            COMMISSION, profit, status
        ])
    return rows

def parse_wconcept(filepath):
    """
    W컨셉 상품준비중내역 - 구버전(31컬럼) / 신버전(32컬럼) 자동 감지
    구버전: [10]상품코드 [11]상품명 [12]옵션1 [13]옵션2 [15]수량 [18]판매가 [23]취소사유 [29]스타일코드
    신버전: [10]대표상품코드 [11]상품코드 [12]상품명 [13]옵션1 [14]옵션2 [16]수량 [19]판매가 [24]취소사유 [30]스타일코드
    """
    print(f"  📦 W컨셉 파싱: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath)
    ws_f = wb.active
    all_rows = list(ws_f.iter_rows(values_only=True))
    if not all_rows:
        return []

    # 헤더로 버전 감지
    header = all_rows[0]
    new_fmt = len(header) >= 32 and str(header[10]).strip() == "대표상품코드"
    if new_fmt:
        # 신버전 (32컬럼)
        i_name, i_color, i_size, i_qty, i_price, i_cancel, i_code, i_base = 12, 13, 14, 16, 19, 24, 30, 10
    else:
        # 구버전 (31컬럼)
        i_name, i_color, i_size, i_qty, i_price, i_cancel, i_code, i_base = 11, 12, 13, 15, 18, 23, 29, 10
    # 주문일 컬럼 자동 탐지 (없으면 index 1 사용 — W컨셉 첫번째 컬럼은 주문번호)
    i_date_wc = find_col(header, "주문일", "주문일시", "결제일") 
    if i_date_wc is None:
        i_date_wc = 1

    rows = []
    for i, row in enumerate(all_rows):
        if i == 0: continue
        if not row[i_name]: continue
        color  = str(row[i_color]) if row[i_color] else "-"
        size   = str(row[i_size])  if row[i_size]  else "-"
        try:
            qty = int(row[i_qty]) if row[i_qty] else 1
        except (ValueError, TypeError):
            qty = 1
        try:
            price = int(row[i_price]) if row[i_price] else 0
        except (ValueError, TypeError):
            price = 0
        total  = price * qty
        profit = round(total * (1 - COMMISSION / 100))
        status = "취소" if row[i_cancel] else "정상"
        code   = str(row[i_code]) if row[i_code] else str(row[i_base])
        # 주문일: 헤더에서 자동 탐지 (파일마다 위치 다를 수 있음)
        rows.append([
            "W컨셉", fmt_date(row[i_date_wc]),
            str(row[i_name]), code,
            color, size, qty, total,
            COMMISSION, profit, status
        ])
    return rows

def parse_ssf(filepath):
    print(f"  📦 SSF 파싱: {os.path.basename(filepath)}")
    wb = xlrd.open_workbook(filepath)
    ws_f = wb.sheet_by_index(0)
    if ws_f.nrows < 2:
        return []
    # 헤더에서 주문일 컬럼 자동 탐지
    header = ws_f.row_values(0)
    i_date = next((i for i, h in enumerate(header) if h and "주문일" in str(h)), 0)
    rows = []
    for i in range(1, ws_f.nrows):
        row = ws_f.row_values(i)
        if not row[21]: continue
        color, size = parse_color_size_slash(row[22])
        qty    = int(row[23]) if row[23] else 1
        price  = int(row[28]) if row[28] else 0
        comm   = int(row[29]) if row[29] else COMMISSION
        total  = price * qty
        profit = round(total * (1 - comm / 100))
        status = str(row[25]) if row[25] else ""
        rows.append([
            "SSF", fmt_date(row[i_date]),
            str(row[21]), str(row[20]),
            color, size, qty, total,
            comm, profit, status
        ])
    return rows

def parse_sivillage(filepath):
    """
    SI Village 발송처리목록 파싱
    컬럼 매핑:
      [2]  배송진행상태 → 주문상태
      [7]  업체상품번호 → 상품코드
      [8]  상품명
      [9]  상품옵션(SKU) → 컬러/사이즈 (예: '더스트핑크/FREE')
      [10] 수량
      [11] 판매금액 → 판매가 (총액)
      [18] 결제완료일시 → 주문일 (예: '20260525061821')
    """
    print(f"  📦 SI Village 파싱: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath)
    ws_f = wb.active
    rows = []
    for i, row in enumerate(ws_f.iter_rows(values_only=True)):
        if i == 0: continue          # 헤더 스킵
        if not row[8]: continue      # 상품명 없으면 스킵

        color, size = parse_color_size_slash(row[9])
        qty    = int(row[10]) if row[10] else 1
        total  = int(row[11]) if row[11] else 0   # 판매금액 (총액)
        profit = round(total * (1 - COMMISSION / 100))
        status = str(row[2]) if row[2] else "정상"
        code   = str(row[7]) if row[7] else "-"
        date   = fmt_date_sivillage(row[18])

        rows.append([
            "SI Village", date,
            str(row[8]), code,
            color, size, qty, total,
            COMMISSION, profit, status
        ])
    return rows


def parse_color_size_cafe24(option_name):
    """Cafe24 옵션: '상품명(색상=워시드차콜, 사이즈=M)' 또는 '상품명(사이즈=L)' 형식"""
    color, size = "-", "-"
    if not option_name:
        return color, size
    s = str(option_name)
    m_color = re.search(r'색상=([^,)]+)', s)
    m_size  = re.search(r'사이즈=([^,)]+)', s)
    if m_color:
        color = m_color.group(1).strip()
    if m_size:
        size  = m_size.group(1).strip()
    # 색상 없고 괄호 안에 단어만 있는 경우: 상품명 (차콜그레이)(사이즈=L)
    if color == "-" and not m_size:
        m_paren = re.search(r'\(([^)]+)\)', s)
        if m_paren:
            val = m_paren.group(1).strip()
            if not val.startswith('사이즈=') and not val.startswith('색상='):
                color = val
    return color, size

def parse_cafe24(filepath):
    """Cafe24 자사몰 주문 엑셀 파싱 (최근 주문내역.xlsx / 결제완료내역.xlsx)"""
    print(f"  📦 Cafe24 파싱: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath)
    ws_f = wb.active
    rows = []
    for i, row in enumerate(ws_f.iter_rows(values_only=True)):
        if i == 0:
            continue
        order_no = str(row[2]) if row[2] else ""
        product  = str(row[8]) if row[8] else ""
        if not product:
            continue
        # 날짜: 발주일(index 19) 우선, 없으면 주문번호 앞 8자리
        if row[19]:
            date = fmt_date(row[19])
        elif len(order_no) >= 8 and order_no[:8].isdigit():
            try:
                date = datetime.strptime(order_no[:8], "%Y%m%d").strftime("%Y-%m-%d")
            except ValueError:
                date = ""
        else:
            date = ""
        color, size = parse_color_size_cafe24(row[9])
        qty    = int(row[10]) if row[10] else 1
        price  = int(row[11]) if row[11] else 0
        total  = price * qty
        profit = round(total * (1 - COMMISSION_CAFE24 / 100))
        # 총 결제금액(index 6) == 0 이면 미결제/취소
        paid   = row[6] if row[6] is not None else 0
        status = "취소" if int(paid) == 0 and price > 0 else "정상"
        code   = str(row[3]) if row[3] else order_no  # 품목별 주문번호
        rows.append([
            "Cafe24", date,
            product, code,
            color, size, qty, total,
            COMMISSION_CAFE24, profit, status
        ])
    return rows
# ── 플랫폼 감지 ─────────────────────────────────────────────────

def parse_coupang(filepath):
    """쿠팡 DeliveryList(날짜)_(N).xlsx — 수수료 11.55% (VAT 포함)
    [9]주문일 [10]등록상품명 [11]등록옵션명("그린 L") [16]업체상품코드 [22]구매수 [23]옵션판매가"""
    COMMISSION_COUPANG = 11.55
    print(f"  📦 쿠팡 파싱: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath)
    ws_f = wb.active
    all_rows = list(ws_f.iter_rows(values_only=True))
    if not all_rows:
        return []
    header = all_rows[0]
    i_date  = find_col(header, "주문일") or 9
    i_prod  = find_col(header, "등록상품명") or 10
    i_opt   = find_col(header, "등록옵션명") or 11
    i_code  = find_col(header, "업체상품코드") or 16
    i_qty   = find_col(header, "구매수(수량)", "구매수") or 22
    i_price = find_col(header, "옵션판매가(판매단가)", "옵션판매가") or 23
    SIZE_PAT = re.compile(r'^(XXS|XS|S|M|L|XL|XXL|2XL|3XL|FREE|\d{2,3})$', re.I)
    rows = []
    for i, row in enumerate(all_rows):
        if i == 0 or not row[i_prod]:
            continue
        # 옵션 "그린 L" → 컬러/사이즈 분리 (마지막 토큰이 사이즈 패턴이면 사이즈)
        color, size = "-", "-"
        opt = str(row[i_opt] or "").strip()
        if opt:
            parts = opt.split()
            if parts and SIZE_PAT.match(parts[-1]):
                size = parts[-1]
                color = " ".join(parts[:-1]) or "-"
            else:
                color = opt
        qty    = int(row[i_qty]) if row[i_qty] else 1
        price  = int(float(row[i_price])) if row[i_price] else 0
        total  = price * qty
        profit = round(total * (1 - COMMISSION_COUPANG / 100))
        rows.append([
            "쿠팡", fmt_date(row[i_date]),
            str(row[i_prod]), str(row[i_code] or "-"),
            color, size, qty, total,
            COMMISSION_COUPANG, profit, "결제완료"
        ])
    return rows

def detect_platform(filepath):
    name = os.path.basename(filepath).lower()
    if "29cm" in name:        return "29cm"
    if "상품준비중" in name:   return "wconcept"
    if "배송조회" in name:     return "ssf"
    if "발송처리목록" in name: return "sivillage"
    if "deliverylist" in name: return "coupang"
    if "주문내역" in name:       return "cafe24"
    if "결제완료내역" in name:    return "cafe24"
    return None

# ── 메인 ────────────────────────────────────────────────────────

def main(files):
    ensure_header()
    existing = get_existing_keys()
    all_new = []

    for filepath in files:
        platform = detect_platform(filepath)
        if not platform:
            print(f"  ⚠️ 플랫폼 감지 실패: {filepath}")
            continue

        if platform == "29cm":         rows = parse_29cm(filepath)
        elif platform == "wconcept":   rows = parse_wconcept(filepath)
        elif platform == "ssf":        rows = parse_ssf(filepath)
        elif platform == "sivillage":  rows = parse_sivillage(filepath)
        elif platform == "coupang":    rows = parse_coupang(filepath)
        elif platform == "cafe24":       rows = parse_cafe24(filepath)
        else:
            rows = []

        added = 0
        # 파일 내 동일 키 카운트 (같은 파일에서 동일 상품 여러 건 처리)
        file_key_count = {}
        for row in rows:
            key = f"{row[0]}|{row[1]}|{row[3]}|{row[4]}|{row[5]}|{row[6]}"
            file_key_count[key] = file_key_count.get(key, 0) + 1
            # 시트에 이미 같은 키가 file_key_count[key]번 이상 있으면 중복
            existing_count = existing.get(key, 0)
            if file_key_count[key] > existing_count:
                all_new.append(row)
                existing[key] = existing.get(key, 0) + 1
                added += 1
        print(f"  ✅ {added}건 추가 ({len(rows)-added}건 중복 스킵)")

    if all_new:
        ws.append_rows(all_new, value_input_option="USER_ENTERED")
        print(f"\n🎉 총 {len(all_new)}건 시트에 저장 완료!")
    else:
        print("\n새로 추가할 데이터 없음.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python3 platform_to_sheets.py [파일경로1] [파일경로2] ...")
        print("지원 플랫폼:")
        print("  29CM       → 파일명에 '29cm' 포함")
        print("  W컨셉      → 파일명에 '상품준비중' 포함")
        print("  SSF        → 파일명에 '배송조회' 포함")
        print("  SI Village → 파일명에 '발송처리목록' 포함")
        print("  Cafe24     → 파일명에 '주문내역' 또는 '결제완료내역' 포함")
    else:
        main(sys.argv[1:])
